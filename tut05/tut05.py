import csv        # importing csv library
import os         # importing os library
import openpyxl   # importing openpyxl library
os.system("cls")
try:                               # using try and except block to create output directory
    os.mkdir("./output/")
except:
    print("Output directory already exsists.")

def marksheet_overall():           #function to create required overall sheet in the workbook named after each roll no.
    to_write={}                    #dictionary to store data to be input into overall sheet
    header=[]                      #list to store the values of the 1st row of names-roll.csv file
    headfile=[]                    #list to store the values of the 1st row of grades.csv file
    with open("names-roll.csv","r")as file:                                #reading names-roll.csv file
        reader=csv.reader(file)
        counter=0
        for row in reader:                                                 #iterating through each row of names-roll.csv file
            if counter==0:
                header=row                                                 #header list stores the 1st row of names-roll.csv
                counter=counter+1
                continue
            if row[0] not in to_write:                                     #checking for if the roll no. already exists
                to_write[row[0]]=[[header[0]+" No.",row[0]]]               #if not present then appending the header list values in to_write dictionary
            to_write[row[0]].append([header[1]+" of the Student",row[1]])  #appending the 2nd row values into to_write dictionary
            string= str(row[0])                                            #converting the roll no. into a string type
            disp=string[4:6]                                               #extracting the branch type from roll no. string
            to_write[row[0]].append(["Discipline",disp])                   #appending 3rd row values for corresponding roll no.
            to_write[row[0]].append(["Semester No.",1,2,3,4,5,6,7,8])      #appending 4th row values for corresponding roll no.
            with open("grades.csv","r")as f:       #reading grades.csv file
                reader=csv.reader(f)
                count=0
                sum_credit_sem1=0                  #initializing variables for each semester to store the sum of all credits in each sem
                sum_credit_sem2=0
                sum_credit_sem3=0
                sum_credit_sem4=0
                sum_credit_sem5=0
                sum_credit_sem6=0
                sum_credit_sem7=0
                sum_credit_sem8=0
                sum_crd_grd1=0                     #initializing variables to store the sum of all products of credits and corresponding grade numeric equivalent in each sem
                sum_crd_grd2=0
                sum_crd_grd3=0
                sum_crd_grd4=0
                sum_crd_grd5=0
                sum_crd_grd6=0
                sum_crd_grd7=0
                sum_crd_grd8=0
                spi_sem1=0                         #initializing variables to store SPI for each sem
                spi_sem2=0
                spi_sem3=0
                spi_sem4=0
                spi_sem5=0
                spi_sem6=0
                spi_sem7=0
                spi_sem8=0
                tot_crd1=0                         #initializing variables to store the cumulative sum of all credits at the end of each sem
                tot_crd2=0
                tot_crd3=0
                tot_crd4=0
                tot_crd5=0
                tot_crd6=0
                tot_crd7=0
                tot_crd8=0
                cpi1=0                            #initializing variables to store CPI for each sem
                cpi2=0
                cpi3=0
                cpi4=0
                cpi5=0
                cpi6=0
                cpi7=0
                cpi8=0
                for grades_row in reader:          #iterating through each row in grades.csv file
                    if count==0:
                        headfile=grades_row        #storing the 1st row in headfile list
                        count=count+1
                        continue
                    if grades_row[4]=="AA":        #converting each grade into its grade numerinc equivalent
                        grades_row[4]="10"
                    if grades_row[4]=="AB":
                        grades_row[4]="9"
                    if grades_row[4]=="BB" or grades_row[4]==' BB':
                        grades_row[4]="8"
                    if grades_row[4]=="BC":
                        grades_row[4]="7"
                    if grades_row[4]=="CC":
                        grades_row[4]="6"
                    if grades_row[4]=="CD":
                        grades_row[4]="5"
                    if grades_row[4]=="DD" or grades_row[4]=="DD*":
                        grades_row[4]="4"
                    if grades_row[4]=="F" or grades_row[4]=="F*":
                        grades_row[4]="0"
                    if grades_row[4]=="I":
                        grades_row[4]="0"
                    if grades_row[0]==row[0]:                      #checking if the roll no. stored in the row of grades.csv file matches with the roll no. stored in the row of names-roll.csv file
                        if grades_row[1]=="1":                     #checking for semester 1
                            sum_credit_sem1=sum_credit_sem1+int(grades_row[3])                     #storing the total credits in sem1             
                            sum_crd_grd1=sum_crd_grd1+(int(grades_row[3])*int(grades_row[4]))      #storing the sum total of credits*grade_numeric_equivalent in sem1
                        if grades_row[1]=="2":                     #checking for semester 2
                            sum_credit_sem2=sum_credit_sem2+int(grades_row[3])                     #storing the total credits in sem2
                            sum_crd_grd2=sum_crd_grd2+(int(grades_row[3])*int(grades_row[4]))      #storing the sum total of credits*grade_numeric_equivalent in sem2
                        if grades_row[1]=="3":                     #checking for semester 3
                            sum_credit_sem3=sum_credit_sem3+int(grades_row[3])                     #storing the total credits in sem3
                            sum_crd_grd3=sum_crd_grd3+(int(grades_row[3])*int(grades_row[4]))      #storing the sum total of credits*grade_numeric_equivalent in sem3
                        if grades_row[1]=="4":                     #checking for semester 4
                            sum_credit_sem4=sum_credit_sem4+int(grades_row[3])                     #storing the total credits in sem4
                            sum_crd_grd4=sum_crd_grd4+(int(grades_row[3])*int(grades_row[4]))      #storing the sum total of credits*grade_numeric_equivalent in sem4
                        if grades_row[1]=="5":                     #checking for semester 5
                            sum_credit_sem5=sum_credit_sem5+int(grades_row[3])                     #storing the total credits in sem5
                            sum_crd_grd5=sum_crd_grd5+(int(grades_row[3])*int(grades_row[4]))      #storing the sum total of credits*grade_numeric_equivalent in sem5
                        if grades_row[1]=="6":                     #checking for semester 6
                            sum_credit_sem6=sum_credit_sem6+int(grades_row[3])                     #storing the total credits in sem6
                            sum_crd_grd6=sum_crd_grd6+(int(grades_row[3])*int(grades_row[4]))      #storing the sum total of credits*grade_numeric_equivalent in sem6
                        if grades_row[1]=="7":                     #checking for semester 7
                            sum_credit_sem7=sum_credit_sem7+int(grades_row[3])                     #storing the total credits in sem7
                            sum_crd_grd7=sum_crd_grd7+(int(grades_row[3])*int(grades_row[4]))      #storing the sum total of credits*grade_numeric_equivalent in sem7
                        if grades_row[1]=="8":                     #checking for semester 8
                            sum_credit_sem8=sum_credit_sem8+int(grades_row[3])                     #storing the total credits in sem8
                            sum_crd_grd8=sum_crd_grd8+(int(grades_row[3])*int(grades_row[4]))      #storing the sum total of credits*grade_numeric_equivalent in sem8
                if sum_credit_sem1!=0:
                    spi_sem1=sum_crd_grd1/sum_credit_sem1        #calculating SPI for sem1
                else:
                    spi_sem1=0 
                if sum_credit_sem2!=0:
                    spi_sem2=sum_crd_grd2/sum_credit_sem2        #calculating SPI for sem2
                else:
                    spi_sem2=0  
                if sum_credit_sem3!=0:
                    spi_sem3=sum_crd_grd3/sum_credit_sem3        #calculating SPI for sem3
                else:
                    spi_sem3=0 
                if sum_credit_sem4!=0:
                    spi_sem4=sum_crd_grd4/sum_credit_sem4        #calculating SPI for sem4
                else:
                    spi_sem4=0 
                if sum_credit_sem5!=0:
                    spi_sem5=sum_crd_grd5/sum_credit_sem5        #calculating SPI for sem5
                else:
                    spi_sem5=0 
                if sum_credit_sem6!=0:
                    spi_sem6=sum_crd_grd6/sum_credit_sem6        #calculating SPI for sem6
                else:
                    spi_sem6=0 
                if sum_credit_sem7!=0:
                    spi_sem7=sum_crd_grd7/sum_credit_sem7        #calculating SPI for sem7
                else:
                    spi_sem7=0 
                if sum_credit_sem8!=0:
                    spi_sem8=sum_crd_grd8/sum_credit_sem8        #calculating SPI for sem8
                else:
                    spi_sem8=0 
                tot_crd1=sum_credit_sem1                         #storing the cumulative sum of total credits at end of each semester
                tot_crd2=sum_credit_sem1+sum_credit_sem2
                tot_crd3=sum_credit_sem1+sum_credit_sem2+sum_credit_sem3
                tot_crd4=sum_credit_sem1+sum_credit_sem2+sum_credit_sem3+sum_credit_sem4
                tot_crd5=sum_credit_sem1+sum_credit_sem2+sum_credit_sem3+sum_credit_sem4+sum_credit_sem5
                tot_crd6=sum_credit_sem1+sum_credit_sem2+sum_credit_sem3+sum_credit_sem4+sum_credit_sem5+sum_credit_sem6
                tot_crd7=sum_credit_sem1+sum_credit_sem2+sum_credit_sem3+sum_credit_sem4+sum_credit_sem5+sum_credit_sem6+sum_credit_sem7
                tot_crd8=sum_credit_sem1+sum_credit_sem2+sum_credit_sem3+sum_credit_sem4+sum_credit_sem5+sum_credit_sem6+sum_credit_sem7+sum_credit_sem8
                cpi1=spi_sem1                                    #calculating CPI for each semester
                cpi2=((spi_sem1*sum_credit_sem1)+(spi_sem2*sum_credit_sem2))/tot_crd2
                cpi3=((spi_sem1*sum_credit_sem1)+(spi_sem2*sum_credit_sem2)+(spi_sem3*sum_credit_sem3))/tot_crd3
                cpi4=((spi_sem1*sum_credit_sem1)+(spi_sem2*sum_credit_sem2)+(spi_sem3*sum_credit_sem3)+(spi_sem4*sum_credit_sem4))/tot_crd4
                cpi5=((spi_sem1*sum_credit_sem1)+(spi_sem2*sum_credit_sem2)+(spi_sem3*sum_credit_sem3)+(spi_sem4*sum_credit_sem4)+(spi_sem5*sum_credit_sem5))/tot_crd5
                cpi6=((spi_sem1*sum_credit_sem1)+(spi_sem2*sum_credit_sem2)+(spi_sem3*sum_credit_sem3)+(spi_sem4*sum_credit_sem4)+(spi_sem5*sum_credit_sem5)+(spi_sem6*sum_credit_sem6))/tot_crd6
                cpi7=((spi_sem1*sum_credit_sem1)+(spi_sem2*sum_credit_sem2)+(spi_sem3*sum_credit_sem3)+(spi_sem4*sum_credit_sem4)+(spi_sem5*sum_credit_sem5)+(spi_sem6*sum_credit_sem6)+(spi_sem7*sum_credit_sem7))/tot_crd7
                cpi8=((spi_sem1*sum_credit_sem1)+(spi_sem2*sum_credit_sem2)+(spi_sem3*sum_credit_sem3)+(spi_sem4*sum_credit_sem4)+(spi_sem5*sum_credit_sem5)+(spi_sem6*sum_credit_sem6)+(spi_sem7*sum_credit_sem7)+(spi_sem8*sum_credit_sem8))/tot_crd8
            to_write[row[0]].append(["Semester wise Credit Taken",sum_credit_sem1,sum_credit_sem2,sum_credit_sem3,sum_credit_sem4,sum_credit_sem5,sum_credit_sem6,sum_credit_sem7,sum_credit_sem8])                                                                   #appending the values for 5th row for correponding roll no
            to_write[row[0]].append(["SPI",'{0:.2f}'.format(spi_sem1),'{0:.2f}'.format(spi_sem2),'{0:.2f}'.format(spi_sem3),'{0:.2f}'.format(spi_sem4),'{0:.2f}'.format(spi_sem5),'{0:.2f}'.format(spi_sem6),'{0:.2f}'.format(spi_sem7),'{0:.2f}'.format(spi_sem8)])  #appending the values for 6th row for correponding roll no. 
            to_write[row[0]].append(["Total Credits Taken",tot_crd1,tot_crd2,tot_crd3,tot_crd4,tot_crd5,tot_crd6,tot_crd7,tot_crd8])                                                                                                                                  #appending the values for 7th row for correponding roll no
            to_write[row[0]].append(["CPI",'{0:.2f}'.format(cpi1),'{0:.2f}'.format(cpi2),'{0:.2f}'.format(cpi3),'{0:.2f}'.format(cpi4),'{0:.2f}'.format(cpi5),'{0:.2f}'.format(cpi6),'{0:.2f}'.format(cpi7),'{0:.2f}'.format(cpi8)])                                  #appending the values for 8th row for correponding roll no
    for key, value in to_write.items():                                 #iterating using keys and values in to_write dictionary
        wb = openpyxl.Workbook()                                        #creating new workbook
        sheet = wb.active                                               #making sheet active
        for row_num in range(1, 4):                                     
            row = value[row_num-1]
            for col_num in range(1, 3):
                sheet.cell(row=(row_num),                   
                           column=col_num).value = row[col_num-1]       #writing data from the dictionary to_write into the overall sheet
        for row_num in range(4,len(value)+1):
            row = value[row_num-1]
            for col_num in range(1, 10):
                sheet.cell(row=(row_num),                  
                           column=col_num).value = row[col_num-1]       #writing data from the dictionary to_write into the overall sheet
        sheet.title="Overall"                                           #renaming sheet as overall
        wb.save("./output/"+key+".xlsx")                                #saving the xlsx file with name as individual roll no.s from key


def marksheet_Sem1():                       #function to create sheet named sem1 in the already created workbook
    to_write={}                             #dictionary to store data to be written into the sheet
    header=[]                               #list to store values of the 1st row of grades.csv file
    headfile=[]                             #list to store values of the 1st row of subjects_master.csv file
    with open("grades.csv","r")as file:     #reading grades.csv file
        reader=csv.reader(file)
        counter=0                           
        Slno_count=1                        #initializing the variable that store serial no. of the data in sheet
        for row in reader:                  #iterating through each row of grades.csv file
            if counter==0:
                header=row                  #header list stores the 1st row of grades.csv file
                counter=counter+1
                continue
            if row[0] not in to_write:      #checking for if the roll no. already exists
                to_write[row[0]]=[["Sl No.","Subject No.","Subject Name","L-T-P",header[3],"Subject Type",header[4]]]     #if not present then appending the header list values in to_write dictionary
                Slno_count=1                #again setting slno_count variable to 1 so that serial no. is stored starting from 1 for every different roll no.
            if row[1]=='1':                 #checking for sem1
                with open("subjects_master.csv","r")as f:           #reading subjects_master.csv file
                    reader=csv.reader(f)
                    count=0
                    for sub_row in reader:                          #iterating through each row of subjects_master.csv file
                        if count==0:
                            headfile=sub_row                        #headfile list stores the 1st row of subjects_master.csv file
                            count=count+1
                            continue                        
                        if sub_row[0]==row[2]:                      #checking if subno. stored in the row of subjects_master.csv file matches with the subno. stored in the row of grades.csv file
                            to_write[row[0]].append([Slno_count,row[2],sub_row[1],sub_row[2],row[3],row[5],row[4]])     #if above statement is true then appending required values from both csv files into to_write dictionary
                            Slno_count=Slno_count+1                 #increasing variable value by 1 for serial no.

    for key, value in to_write.items():                                #iterating using keys and values in to_write dictionary
        wb = openpyxl.load_workbook("./output/"+key+".xlsx")           #loading already created workbook corresponding to individual roll no.s
        sheetname="Sem1"                                               
        sheet=wb.create_sheet(sheetname)                               #creating new worksheet named as Sem1
        for row_num in range(1, len(value)+1):
            row = value[row_num-1]
            for col_num in range(1, 8):
                sheet.cell(row=(row_num),                  
                            column=col_num).value = row[col_num-1]     #writing data from the dictionary to_write into Sem1 sheet
        wb.save("./output/"+key+".xlsx")                               #saving changes made into the workbook

#All other functions for different semester sheets are similar to Sem1 sheet

def marksheet_Sem2():                            #function to create sheet named sem2 in the already created workbook
    to_write={}                                  #dictionary to store data to be written into the sheet
    header=[]                                    #list to store values of the 1st row of grades.csv file
    headfile=[]                                  #list to store values of the 1st row of subjects_master.csv file
    with open("grades.csv","r")as file:          #reading grades.csv file
        reader=csv.reader(file)
        counter=0
        Slno_count=1                             #initializing the variable that store serial no. of the data in sheet
        for row in reader:                       #iterating through each row of grades.csv file
            if counter==0:
                header=row                       #header list stores the 1st row of grades.csv file
                counter=counter+1
                continue
            if row[0] not in to_write:           #checking for if the roll no. already exists
                to_write[row[0]]=[["Sl No.","Subject No.","Subject Name","L-T-P",header[3],"Subject Type",header[4]]]    #if not present then appending the header list values in to_write dictionary
                Slno_count=1                     #again setting slno_count variable to 1 so that serial no. is stored starting from 1 for every different roll no.
            if row[1]=='2':                      #checking for sem2
                with open("subjects_master.csv","r")as f:                      #reading subjects_master.csv file
                    reader=csv.reader(f)
                    count=0
                    for sub_row in reader:                                     #iterating through each row of subjects_master.csv file
                        if count==0:
                            headfile=sub_row                                   #headfile list stores the 1st row of subjects_master.csv file
                            count=count+1
                            continue                        
                        if sub_row[0]==row[2]:                                 #checking if subno. stored in the row of subjects_master.csv file matches with the subno. stored in the row of grades.csv file
                            to_write[row[0]].append([Slno_count,row[2],sub_row[1],sub_row[2],row[3],row[5],row[4]])         #if above statement is true then appending required values from both csv files into to_write dictionary
                            Slno_count=Slno_count+1                            #increasing variable value by 1 for serial no.

    for key, value in to_write.items():                                        #iterating using keys and values in to_write dictionary
        wb = openpyxl.load_workbook("./output/"+key+".xlsx")                   #loading already created workbook corresponding to individual roll no.s
        sheetname="Sem2"
        sheet=wb.create_sheet(sheetname)                                       #creating new worksheet named as Sem2
        for row_num in range(1, len(value)+1):
            row = value[row_num-1]
            for col_num in range(1, 8):
                sheet.cell(row=(row_num),                  
                            column=col_num).value = row[col_num-1]             #writing data from the dictionary to_write into Sem2 sheet
        wb.save("./output/"+key+".xlsx")                                       #saving changes made into the workbook

def marksheet_Sem3():                               #function to create sheet named sem3 in the already created workbook
    to_write={}                                     #dictionary to store data to be written into the sheet
    header=[]                                       #list to store values of the 1st row of grades.csv file
    headfile=[]                                     #list to store values of the 1st row of subjects_master.csv file
    with open("grades.csv","r")as file:             #reading grades.csv file
        reader=csv.reader(file)
        counter=0
        Slno_count=1                                #initializing the variable that store serial no. of the data in sheet
        for row in reader:                          #iterating through each row of grades.csv file
            if counter==0:
                header=row                          #header list stores the 1st row of grades.csv file
                counter=counter+1
                continue
            if row[0] not in to_write:              #checking for if the roll no. already exists
                to_write[row[0]]=[["Sl No.","Subject No.","Subject Name","L-T-P",header[3],"Subject Type",header[4]]]       #if not present then appending the header list values in to_write dictionary
                Slno_count=1                        #again setting slno_count variable to 1 so that serial no. is stored starting from 1 for every different roll no.
            if row[1]=='3':                         #checking for sem3
                with open("subjects_master.csv","r")as f:                       #reading subjects_master.csv file
                    reader=csv.reader(f)
                    count=0
                    for sub_row in reader:                                      #iterating through each row of subjects_master.csv file
                        if count==0:
                            headfile=sub_row                                    #headfile list stores the 1st row of subjects_master.csv file
                            count=count+1
                            continue                        
                        if sub_row[0]==row[2]:                                  #checking if subno. stored in the row of subjects_master.csv file matches with the subno. stored in the row of grades.csv file
                            to_write[row[0]].append([Slno_count,row[2],sub_row[1],sub_row[2],row[3],row[5],row[4]])         #if above statement is true then appending required values from both csv files into to_write dictionary
                            Slno_count=Slno_count+1                             #increasing variable value by 1 for serial no.

    for key, value in to_write.items():                                  #iterating using keys and values in to_write dictionary
        wb = openpyxl.load_workbook("./output/"+key+".xlsx")             #loading already created workbook corresponding to individual roll no.s
        sheetname="Sem3"
        sheet=wb.create_sheet(sheetname)                                 #creating new worksheet named as Sem3
        for row_num in range(1, len(value)+1):
            row = value[row_num-1]
            for col_num in range(1, 8):
                sheet.cell(row=(row_num),                  
                            column=col_num).value = row[col_num-1]       #writing data from the dictionary to_write into Sem3 sheet
        wb.save("./output/"+key+".xlsx")                                 #saving changes made into the workbook


def marksheet_Sem4():                                          #function to create sheet named sem4 in the already created workbook
    to_write={}                                                #dictionary to store data to be written into the sheet
    header=[]                                                  #list to store values of the 1st row of grades.csv file
    headfile=[]                                                #list to store values of the 1st row of subjects_master.csv file
    with open("grades.csv","r")as file:                        #reading grades.csv file
        reader=csv.reader(file)
        counter=0
        Slno_count=1                                           #initializing the variable that store serial no. of the data in sheet
        for row in reader:                                     #iterating through each row of grades.csv file
            if counter==0:
                header=row                                     #header list stores the 1st row of grades.csv file
                counter=counter+1
                continue
            if row[0] not in to_write:                         #checking for if the roll no. already exists
                to_write[row[0]]=[["Sl No.","Subject No.","Subject Name","L-T-P",header[3],"Subject Type",header[4]]]        #if not present then appending the header list values in to_write dictionary
                Slno_count=1                                   #again setting slno_count variable to 1 so that serial no. is stored starting from 1 for every different roll no.
            if row[1]=='4':                                    #checking for sem4
                with open("subjects_master.csv","r")as f:      #reading subjects_master.csv file
                    reader=csv.reader(f)
                    count=0
                    for sub_row in reader:                     #iterating through each row of subjects_master.csv file
                        if count==0:
                            headfile=sub_row                   #headfile list stores the 1st row of subjects_master.csv file
                            count=count+1
                            continue                        
                        if sub_row[0]==row[2]:                 #checking if subno. stored in the row of subjects_master.csv file matches with the subno. stored in the row of grades.csv file
                            to_write[row[0]].append([Slno_count,row[2],sub_row[1],sub_row[2],row[3],row[5],row[4]])            #if above statement is true then appending required values from both csv files into to_write dictionary
                            Slno_count=Slno_count+1            #increasing variable value by 1 for serial no.

    for key, value in to_write.items():                        #iterating using keys and values in to_write dictionary
        wb = openpyxl.load_workbook("./output/"+key+".xlsx")   #loading already created workbook corresponding to individual roll no.s
        sheetname="Sem4"       
        sheet=wb.create_sheet(sheetname)                       #creating new worksheet named as Sem4
        for row_num in range(1, len(value)+1):
            row = value[row_num-1]
            for col_num in range(1, 8):
                sheet.cell(row=(row_num),                  
                            column=col_num).value = row[col_num-1]         #writing data from the dictionary to_write into Sem4 sheet
        wb.save("./output/"+key+".xlsx")                                   #saving changes made into the workbook

def marksheet_Sem5():                                          #function to create sheet named sem5 in the already created workbook
    to_write={}                                                #dictionary to store data to be written into the sheet
    header=[]                                                  #list to store values of the 1st row of grades.csv file
    headfile=[]                                                #list to store values of the 1st row of subjects_master.csv file
    with open("grades.csv","r")as file:                        #reading grades.csv file
        reader=csv.reader(file)                                
        counter=0
        Slno_count=1                                           #initializing the variable that store serial no. of the data in sheet
        for row in reader:                                     #iterating through each row of grades.csv file
            if counter==0:
                header=row                                     #header list stores the 1st row of grades.csv file
                counter=counter+1
                continue
            if row[0] not in to_write:                         #checking for if the roll no. already exists
                to_write[row[0]]=[["Sl No.","Subject No.","Subject Name","L-T-P",header[3],"Subject Type",header[4]]]         #if not present then appending the header list values in to_write dictionary
                Slno_count=1                                   #again setting slno_count variable to 1 so that serial no. is stored starting from 1 for every different roll no.
            if row[1]=='5':                                    #checking for sem5
                with open("subjects_master.csv","r")as f:      #reading subjects_master.csv file
                    reader=csv.reader(f)
                    count=0
                    for sub_row in reader:                     #iterating through each row of subjects_master.csv file
                        if count==0:
                            headfile=sub_row                   #headfile list stores the 1st row of subjects_master.csv file
                            count=count+1
                            continue                        
                        if sub_row[0]==row[2]:                 #checking if subno. stored in the row of subjects_master.csv file matches with the subno. stored in the row of grades.csv file
                            to_write[row[0]].append([Slno_count,row[2],sub_row[1],sub_row[2],row[3],row[5],row[4]])           #if above statement is true then appending required values from both csv files into to_write dictionary
                            Slno_count=Slno_count+1            #increasing variable value by 1 for serial no.

    for key, value in to_write.items():                        #iterating using keys and values in to_write dictionary
        wb = openpyxl.load_workbook("./output/"+key+".xlsx")   #loading already created workbook corresponding to individual roll no.s
        sheetname="Sem5"
        sheet=wb.create_sheet(sheetname)                       #creating new worksheet named as Sem5
        for row_num in range(1, len(value)+1):
            row = value[row_num-1]
            for col_num in range(1, 8):
                sheet.cell(row=(row_num),                  
                            column=col_num).value = row[col_num-1]             #writing data from the dictionary to_write into Sem5 sheet
        wb.save("./output/"+key+".xlsx")                                       #saving changes made into the workbook


def marksheet_Sem6():                                          #function to create sheet named sem6 in the already created workbook
    to_write={}                                                #dictionary to store data to be written into the sheet
    header=[]                                                  #list to store values of the 1st row of grades.csv file
    headfile=[]                                                #list to store values of the 1st row of subjects_master.csv file
    with open("grades.csv","r")as file:                        #reading grades.csv file
        reader=csv.reader(file)
        counter=0
        Slno_count=1                                           #initializing the variable that store serial no. of the data in sheet
        for row in reader:                                     #iterating through each row of grades.csv file
            if counter==0:
                header=row                                     #header list stores the 1st row of grades.csv file
                counter=counter+1
                continue
            if row[0] not in to_write:                         #checking for if the roll no. already exists
                to_write[row[0]]=[["Sl No.","Subject No.","Subject Name","L-T-P",header[3],"Subject Type",header[4]]]            #if not present then appending the header list values in to_write dictionary
                Slno_count=1                                   #again setting slno_count variable to 1 so that serial no. is stored starting from 1 for every different roll no.
            if row[1]=='6':                                    #checking for sem6
                with open("subjects_master.csv","r")as f:      #reading subjects_master.csv file
                    reader=csv.reader(f)
                    count=0
                    for sub_row in reader:                     #iterating through each row of subjects_master.csv file
                        if count==0:
                            headfile=sub_row                   #headfile list stores the 1st row of subjects_master.csv file
                            count=count+1
                            continue                        
                        if sub_row[0]==row[2]:                 #checking if subno. stored in the row of subjects_master.csv file matches with the subno. stored in the row of grades.csv file
                            to_write[row[0]].append([Slno_count,row[2],sub_row[1],sub_row[2],row[3],row[5],row[4]])           #if above statement is true then appending required values from both csv files into to_write dictionary
                            Slno_count=Slno_count+1            #increasing variable value by 1 for serial no.

    for key, value in to_write.items():                        #iterating using keys and values in to_write dictionary
        wb = openpyxl.load_workbook("./output/"+key+".xlsx")   #loading already created workbook corresponding to individual roll no.s
        sheetname="Sem6"
        sheet=wb.create_sheet(sheetname)                       #creating new worksheet named as Sem6
        for row_num in range(1, len(value)+1):
            row = value[row_num-1]
            for col_num in range(1, 8):
                sheet.cell(row=(row_num),                  
                            column=col_num).value = row[col_num-1]           #writing data from the dictionary to_write into Sem6 sheet
        wb.save("./output/"+key+".xlsx")                                     #saving changes made into the workbook


def marksheet_Sem7():                                #function to create sheet named sem7 in the already created workbook
    to_write={}                                      #dictionary to store data to be written into the sheet
    header=[]                                        #list to store values of the 1st row of grades.csv file
    headfile=[]                                      #list to store values of the 1st row of subjects_master.csv file
    with open("grades.csv","r")as file:              #reading grades.csv file
        reader=csv.reader(file)
        counter=0
        Slno_count=1                                 #initializing the variable that store serial no. of the data in sheet
        for row in reader:                           #iterating through each row of grades.csv file
            if counter==0:
                header=row                           #header list stores the 1st row of grades.csv file
                counter=counter+1
                continue
            if row[0] not in to_write:               #checking for if the roll no. already exists
                to_write[row[0]]=[["Sl No.","Subject No.","Subject Name","L-T-P",header[3],"Subject Type",header[4]]]           #if not present then appending the header list values in to_write dictionary
                Slno_count=1                         #again setting slno_count variable to 1 so that serial no. is stored starting from 1 for every different roll no.
            if row[1]=='7':                          #checking for sem7
                with open("subjects_master.csv","r")as f:          #reading subjects_master.csv file
                    reader=csv.reader(f)
                    count=0
                    for sub_row in reader:                         #iterating through each row of subjects_master.csv file
                        if count==0:
                            headfile=sub_row                       #headfile list stores the 1st row of subjects_master.csv file
                            count=count+1
                            continue                        
                        if sub_row[0]==row[2]:                     #checking if subno. stored in the row of subjects_master.csv file matches with the subno. stored in the row of grades.csv file
                            to_write[row[0]].append([Slno_count,row[2],sub_row[1],sub_row[2],row[3],row[5],row[4]])           #if above statement is true then appending required values from both csv files into to_write dictionary
                            Slno_count=Slno_count+1                #increasing variable value by 1 for serial no.

    for key, value in to_write.items():                                #iterating using keys and values in to_write dictionary
        wb = openpyxl.load_workbook("./output/"+key+".xlsx")           #loading already created workbook corresponding to individual roll no.s
        sheetname="Sem7"
        sheet=wb.create_sheet(sheetname)                               #creating new worksheet named as Sem7
        for row_num in range(1, len(value)+1):
            row = value[row_num-1]
            for col_num in range(1, 8):
                sheet.cell(row=(row_num),                  
                            column=col_num).value = row[col_num-1]     #writing data from the dictionary to_write into Sem7 sheet
        wb.save("./output/"+key+".xlsx")                               #saving changes made into the workbook

def marksheet_Sem8():                                   #function to create sheet named sem8 in the already created workbook
    to_write={}                                         #dictionary to store data to be written into the sheet
    header=[]                                           #list to store values of the 1st row of grades.csv file
    headfile=[]                                         #list to store values of the 1st row of subjects_master.csv file
    with open("grades.csv","r")as file:                 #reading grades.csv file
        reader=csv.reader(file)
        counter=0
        Slno_count=1                                    #initializing the variable that store serial no. of the data in sheet
        for row in reader:                              #iterating through each row of grades.csv file
            if counter==0:
                header=row                              #header list stores the 1st row of grades.csv file
                counter=counter+1
                continue
            if row[0] not in to_write:                  #checking for if the roll no. already exists
                to_write[row[0]]=[["Sl No.","Subject No.","Subject Name","L-T-P",header[3],"Subject Type",header[4]]]       #if not present then appending the header list values in to_write dictionary
                Slno_count=1                            #again setting slno_count variable to 1 so that serial no. is stored starting from 1 for every different roll no.
            if row[1]=='8':                             #checking for sem8
                with open("subjects_master.csv","r")as f:                 #reading subjects_master.csv file
                    reader=csv.reader(f)
                    count=0
                    for sub_row in reader:                                #iterating through each row of subjects_master.csv file
                        if count==0:
                            headfile=sub_row                              #headfile list stores the 1st row of subjects_master.csv file
                            count=count+1
                            continue                        
                        if sub_row[0]==row[2]:                            #checking if subno. stored in the row of subjects_master.csv file matches with the subno. stored in the row of grades.csv file
                            to_write[row[0]].append([Slno_count,row[2],sub_row[1],sub_row[2],row[3],row[5],row[4]])       #if above statement is true then appending required values from both csv files into to_write dictionary
                            Slno_count=Slno_count+1                       #increasing variable value by 1 for serial no.

    for key, value in to_write.items():                                #iterating using keys and values in to_write dictionary
        wb = openpyxl.load_workbook("./output/"+key+".xlsx")           #loading already created workbook corresponding to individual roll no.s
        sheetname="Sem8"
        sheet=wb.create_sheet(sheetname)                               #creating new worksheet named as Sem8
        for row_num in range(1, len(value)+1):
            row = value[row_num-1]
            for col_num in range(1, 8):
                sheet.cell(row=(row_num),                  
                            column=col_num).value = row[col_num-1]     #writing data from the dictionary to_write into Sem8 sheet
        wb.save("./output/"+key+".xlsx")                               #saving changes made into the workbook

def marksheet_Sem10():                                     #function to create sheet named sem10 in the already created workbook
    to_write={}                                            #dictionary to store data to be written into the sheet
    header=[]                                              #list to store values of the 1st row of grades.csv file
    headfile=[]                                            #list to store values of the 1st row of subjects_master.csv file
    with open("grades.csv","r")as file:                    #reading grades.csv file
        reader=csv.reader(file)
        counter=0
        chk_sem10_count=0                                  #variable to check for the existence of Sem10 in grades.csv file corresponding to each roll no.
        Slno_count=1                                       #initializing the variable that store serial no. of the data in sheet
        for row in reader:                                 #iterating through each row of grades.csv file
            if counter==0:
                header=row                                 #header list stores the 1st row of grades.csv file
                counter=counter+1
                continue
            if row[1]=='10':                               #checking for sem10
                if row[0] not in to_write:                 #checking for if the roll no. already exists
                    to_write[row[0]]=[["Sl No.","Subject No.","Subject Name","L-T-P",header[3],"Subject Type",header[4]]]            #if not present then appending the header list values in to_write dictionary
                    Slno_count=1                           #again setting slno_count variable to 1 so that serial no. is stored starting from 1 for every different roll no.
                chk_sem10_count=1                          #setting the value to 1 as sem 10 exists for the corresponding roll no.
                with open("subjects_master.csv","r")as f:          #reading subjects_master.csv file
                    reader=csv.reader(f)
                    count=0
                    for sub_row in reader:                         #iterating through each row of subjects_master.csv file
                        if count==0: 
                            headfile=sub_row                       #headfile list stores the 1st row of subjects_master.csv file
                            count=count+1
                            continue                        
                        if sub_row[0]==row[2]:                     #checking if subno. stored in the row of subjects_master.csv file matches with the subno. stored in the row of grades.csv file
                            to_write[row[0]].append([Slno_count,row[2],sub_row[1],sub_row[2],row[3],row[5],row[4]])           #if above statement is true then appending required values from both csv files into to_write dictionary
                            Slno_count=Slno_count+1                #increasing variable value by 1 for serial no.

    for key, value in to_write.items():                                  #iterating using keys and values in to_write dictionary
        if chk_sem10_count==1:                                           #checking if sem10 exists for different roll values
            wb = openpyxl.load_workbook("./output/"+key+".xlsx")         #loading already created workbook corresponding to individual roll no.s
            sheetname="Sem10"
            sheet=wb.create_sheet(sheetname)                             #creating new worksheet named as Sem10
            for row_num in range(1, len(value)+1):
                row = value[row_num-1]
                for col_num in range(1, 8):
                    sheet.cell(row=(row_num),                  
                                column=col_num).value = row[col_num-1]   #writing data from the dictionary to_write into Sem10 sheet
            wb.save("./output/"+key+".xlsx")                             #saving changes made into the workbook

#Calling functions corresponding to different sheets in the workbook

marksheet_overall()
marksheet_Sem1()
marksheet_Sem2()
marksheet_Sem3()
marksheet_Sem4()
marksheet_Sem5()
marksheet_Sem6()
marksheet_Sem7()
marksheet_Sem8()
marksheet_Sem10()