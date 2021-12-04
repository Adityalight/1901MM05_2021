import streamlit as st
import os
import csv
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import Border
from openpyxl.styles import Side
import smtplib
from email.message import EmailMessage
if not os.path.isdir("./sample_input/"):
    os.mkdir("./sample_input/")
def save_uploadedfile(uploadedfile):
    with open(os.path.join("sample_input",uploadedfile.name),"wb") as f:
        f.write(uploadedfile.getbuffer())
    return st.success("Saved File: {} ".format(uploadedfile.name))
def success_Rollwise():
    return st.success("Roll Number wise Marksheets are successfuly created")
def success_concise():
    return st.success("Concise Marksheet is successfuly created")
def send_email():
    return st.success("Emails are sent successfuly")
def main():
    st.title("Marksheet Generator")
    menu = ["Home","Marksheet Generator","About"]
    choice = st.sidebar.selectbox("Menu",menu)
    if choice == "Home":
        st.subheader("Go to Marksheet Generator option from the menu to generate marksheet")
    elif choice == "Marksheet Generator":
        st.write("Please Upload the correct csv file and generate marksheet" )
        data_file = st.file_uploader("Upload master_roll.csv",type=['csv'])
        if data_file is not None:
            file_details = {"Filename":data_file.name,"FileType":data_file.type,"FileSize":data_file.size}
            st.write(file_details)
            df1 = pd.read_csv(data_file)
            st.dataframe(df1)
            save_uploadedfile(data_file)
        data_file = st.file_uploader("Upload responses.csv",type=['csv'])
        if data_file is not None:
            file_details = {"Filename":data_file.name,"FileType":data_file.type,"FileSize":data_file.size}
            st.write(file_details)
            df2 = pd.read_csv(data_file)
            st.dataframe(df2)
            save_uploadedfile(data_file)
        positive_marking = st.number_input("Marks for Correct Answer",value=5.00,step=1.,format="%.2f")
        negative_marking = st.number_input("Marks for wrong Answer",value=-1.00,step=1.,format="%.2f")
        if st.button("Generate Roll Number wise Marksheet"):
            def check_ANSWER_roll():
                with open('sample_input/responses.csv','r') as file:
                    reader=csv.reader(file)
                    flag=0
                    for row in reader:
                        if row[6]=="ANSWER":
                            flag=1
                        else:
                            continue
                    if flag==0:
                        st.write("no roll number with ANSWER is present, Cannot Generate Marksheet!")
                        exit()
                        
            # positive_marking = 5
            # negative_marking = -1
            result_marksheet_data={}
            Correct_answers=[]
            def individual_roll_result():
                with open('sample_input/responses.csv','r') as file:
                    counter=0
                    Header=[]
                    Total_Questions=28
                    reader=csv.reader(file)
                    for row in reader:
                        if row[6]=="ANSWER":
                            for i in range(7,len(row)):
                                Correct_answers.append(row[i])
                            break
                            # print(Correct_answers)
                with open('sample_input/responses.csv','r') as file:
                    reader=csv.reader(file)
                    for row in reader:
                        if counter==0:
                            Header=row
                            counter+=1
                            continue
                        result_marksheet_data[row[6]]={}
                        Right_answer=0
                        Wrong_answer=0
                        Not_attempted=0
                        marks_scored=0
                        response=[]
                        for i in range(7,len(row)):
                            Student_answer=row[i]
                            response.append([Student_answer,Correct_answers[i-7]])
                            if len(Student_answer.strip())==0:
                                Not_attempted+=1
                            elif Student_answer==Correct_answers[i-7]:
                                marks_scored+=positive_marking
                                Right_answer+=1
                            else:
                                marks_scored+=negative_marking
                                Wrong_answer+=1
                        result_marksheet_data[row[6]]["info"]=[row[0],row[1],row[2],row[3],row[4],row[5],str(marks_scored)+"/"+str(positive_marking*Total_Questions),row[6]]              
                        result_marksheet_data[row[6]]["marked_answers"]=response
                        result_marksheet_data[row[6]]["data"]=[["No.",Right_answer,Wrong_answer,Not_attempted,Total_Questions],["Marking",positive_marking,negative_marking,0,""],["Total",Right_answer*positive_marking,Wrong_answer*negative_marking,"",str(marks_scored)+"/"+str(Total_Questions*positive_marking)]]
                        right_ans_str='['+str(Right_answer)+','+str(Wrong_answer)+','+str(Not_attempted)+']'
                        result_marksheet_data[row[6]]["StatusAns"]=[right_ans_str]

                if not os.path.isdir("./marksheets/"):                               
                    os.mkdir("./marksheets/")
                for key, value in result_marksheet_data.items():
                    from openpyxl import Workbook
                    wb = Workbook()
                    sheet = wb.active
                    sheet.title="quiz"
                    img = openpyxl.drawing.image.Image("IITP LOGO.png")
                    img.anchor = 'A1'
                    img.width = 648
                    img.height = 83
                    sheet.add_image(img)
                    sheet.column_dimensions['A'].width=18
                    sheet.column_dimensions['B'].width=18
                    sheet.column_dimensions['C'].width=18
                    sheet.column_dimensions['D'].width=18
                    sheet.column_dimensions['E'].width=18
                    sheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=5)
                    sheet.cell(row=5,column=1,value='Mark Sheet').font=Font(name='Century',size=18,underline='single',color='000000',bold=True)
                    sheet.cell(row=5,column=1).alignment = Alignment(horizontal='center')
                    sheet.cell(row=6,column=1,value='Name:').font=Font(name='Century',size=12,color='000000')
                    sheet.cell(row=6,column=1).alignment=Alignment(horizontal='right')
                    stud_name=value["info"][3]
                    sheet.cell(row=6,column=2,value=stud_name).font=Font(name='Century',size=12,color='000000',bold=True)
                    sheet.cell(row=6,column=4,value='Exam:').font=Font(name='Century',size=12,color='000000')
                    sheet.cell(row=6,column=4).alignment=Alignment(horizontal='right')
                    sheet.cell(row=6,column=5,value='quiz').font=Font(name='Century',size=12,color='000000',bold=True)
                    sheet.cell(row=7,column=1,value='Roll Number:').font=Font(name='Century',size=12,color='000000')
                    sheet.cell(row=7,column=1).alignment=Alignment(horizontal='right')
                    sheet.cell(row=7,column=2,value=key).font=Font(name='Century',size=12,color='000000',bold=True)

                    header_data = ["","Right","Wrong","Not Attempt","Max"]
                    col_no = 1
                    for i in header_data:
                        sheet.cell(row=9,column=col_no,value=i).font=Font(name='Century',size=12,color='000000',bold=True)
                        sheet.cell(row=9,column=col_no).alignment=Alignment(horizontal='center')
                        sheet.cell(row=9,column=col_no,value=i).border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        col_no+=1

                    data=value["data"]
                    row_no = 10
                    for sub_data in data:
                        col_no=1
                        for x in sub_data:
                            if col_no==1:
                                sheet.cell(row=row_no,column=col_no,value=x).font=Font(name='Century',size=12,color='000000',bold=True)
                                sheet.cell(row=row_no,column=col_no).alignment=Alignment(horizontal='center')
                                sheet.cell(row=row_no,column=col_no,value=x).border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                            elif col_no==2:
                                sheet.cell(row=row_no,column=col_no,value=x).font=Font(name='Century',size=12,color='339933')
                                sheet.cell(row=row_no,column=col_no).alignment=Alignment(horizontal='center')
                                sheet.cell(row=row_no,column=col_no,value=x).border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                            elif col_no==3:
                                sheet.cell(row=row_no,column=col_no,value=x).font=Font(name='Century',size=12,color='FF0000')
                                sheet.cell(row=row_no,column=col_no).alignment=Alignment(horizontal='center')
                                sheet.cell(row=row_no,column=col_no,value=x).border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                            else:
                                sheet.cell(row=row_no,column=col_no,value=x).font=Font(name='Century',size=12,color='000000')
                                sheet.cell(row=row_no,column=col_no).alignment=Alignment(horizontal='center')
                                sheet.cell(row=row_no,column=col_no,value=x).border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                            col_no+=1
                        row_no+=1
                    sheet.cell(row=12,column=5).font=Font(name='Century',size=12,color='0000FF') 
                    for i in [1,2,4,5]:
                        col_no=i
                        if i==1 or i==4:
                            response_header = "Student Ans"
                        else:
                            response_header = "Correct Ans"
                        sheet.cell(row=15,column=col_no,value=response_header).font=Font(name='Century',size=12,color='000000',bold=True)
                        sheet.cell(row=15,column=col_no).alignment=Alignment(horizontal='center')
                        sheet.cell(row=15,column=col_no,value=response_header).border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    row_no=16
                    marked_ans=value["marked_answers"]
                    for i in range(0,len(marked_ans)-3):
                        stud_answer=marked_ans[i][0]
                        corr_answer=marked_ans[i][1]
                        if stud_answer==corr_answer:
                            sheet.cell(row=row_no,column=1,value=stud_answer).font=Font(name='Century',size=12,color='339933')
                            sheet.cell(row=row_no,column=1).alignment=Alignment(horizontal='center')
                            sheet.cell(row=row_no,column=1,value=stud_answer).border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        else:
                            sheet.cell(row=row_no,column=1,value=stud_answer).font=Font(name='Century',size=12,color='FF0000')
                            sheet.cell(row=row_no,column=1).alignment=Alignment(horizontal='center')
                            sheet.cell(row=row_no,column=1,value=stud_answer).border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        
                        sheet.cell(row=row_no,column=2,value=stud_answer).font=Font(name='Century',size=12,color='0000FF')
                        sheet.cell(row=row_no,column=2).alignment=Alignment(horizontal='center')
                        sheet.cell(row=row_no,column=2,value=stud_answer).border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        row_no+=1
                    row_no=16
                    for i in range(len(marked_ans)-3,len(marked_ans)):
                        stud_answer=marked_ans[i][0]
                        corr_answer=marked_ans[i][1]
                        if stud_answer==corr_answer:
                            sheet.cell(row=row_no,column=4,value=stud_answer).font=Font(name='Century',size=12,color='339933')
                            sheet.cell(row=row_no,column=4).alignment=Alignment(horizontal='center')
                            sheet.cell(row=row_no,column=4,value=stud_answer).border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        else:
                            sheet.cell(row=row_no,column=4,value=stud_answer).font=Font(name='Century',size=12,color='FF0000')
                            sheet.cell(row=row_no,column=4).alignment=Alignment(horizontal='center')
                            sheet.cell(row=row_no,column=4,value=stud_answer).border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        
                        sheet.cell(row=row_no,column=5,value=stud_answer).font=Font(name='Century',size=12,color='0000FF')
                        sheet.cell(row=row_no,column=5).alignment=Alignment(horizontal='center')
                        sheet.cell(row=row_no,column=5,value=stud_answer).border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        row_no+=1
                    wb.save("./marksheets/"+key+".xlsx")

            absent_roll_nos = {}
            def blank_marksheet():
                with open('sample_input/master_roll.csv','r') as file:
                    counter=0
                    header=[]
                    reader = csv.reader(file)
                    for row in reader:
                        if counter==0:
                            header=row
                            counter+=1
                            continue
                        with open('sample_input/responses.csv','r') as f:
                            count=0
                            headfile=[]
                            reader=csv.reader(f)
                            flag=0
                            for row_response in reader:
                                if count==0:
                                    headfile=row_response
                                    count+=1
                                    continue
                                if row_response[6]==row[0]:
                                    flag=1
                                    break
                            if flag==0:
                                absent_roll_nos[row[0]]=[row[1]]

                from openpyxl import Workbook
                wb=Workbook()
                sheet=wb.active
                sheet.title="quiz"
                for key,value in absent_roll_nos.items():
                    img = openpyxl.drawing.image.Image("IITP LOGO.png")
                    img.anchor = 'A1'
                    img.width = 648
                    img.height = 83
                    sheet.add_image(img)
                    sheet.column_dimensions['A'].width=18
                    sheet.column_dimensions['B'].width=18
                    sheet.column_dimensions['C'].width=18
                    sheet.column_dimensions['D'].width=18
                    sheet.column_dimensions['E'].width=18
                    sheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=5)
                    sheet.cell(row=5,column=1,value='Mark Sheet').font=Font(name='Century',size=18,underline='single',color='000000',bold=True)
                    sheet.cell(row=5,column=1).alignment = Alignment(horizontal='center')
                    sheet.cell(row=6,column=1,value='Name:').font=Font(name='Century',size=12,color='000000')
                    sheet.cell(row=6,column=1).alignment=Alignment(horizontal='right')
                    stud_name=value[0]
                    sheet.cell(row=6,column=2,value=stud_name).font=Font(name='Century',size=12,color='000000',bold=True)
                    sheet.cell(row=6,column=4,value='Exam:').font=Font(name='Century',size=12,color='000000')
                    sheet.cell(row=6,column=4).alignment=Alignment(horizontal='right')
                    sheet.cell(row=6,column=5,value='quiz').font=Font(name='Century',size=12,color='000000',bold=True)
                    sheet.cell(row=7,column=1,value='Roll Number:').font=Font(name='Century',size=12,color='000000')
                    sheet.cell(row=7,column=1).alignment=Alignment(horizontal='right')
                    sheet.cell(row=7,column=2,value=key).font=Font(name='Century',size=12,color='000000',bold=True)

                    sheet.merge_cells(start_row=9, start_column=2, end_row=11, end_column=4)
                    sheet.cell(row=9,column=2,value='ABSENT').font=Font(name='Century',size=20,color='000000',bold=True)
                    sheet.cell(row=9,column=2).alignment = Alignment(horizontal='center')
                    wb.save("./marksheets/"+key+".xlsx")     
            check_ANSWER_roll()
            individual_roll_result()
            blank_marksheet()
            success_Rollwise()
        

        if st.button("Generate concise Marksheet"):
            result_marksheet_data={}
            Correct_answers=[]
            def individual_roll_result():
                with open('sample_input/responses.csv','r') as file:
                    counter=0
                    Header=[]
                    Total_Questions=28
                    reader=csv.reader(file)
                    for row in reader:
                        if row[6]=="ANSWER":
                            for i in range(7,len(row)):
                                Correct_answers.append(row[i])
                            break
                            # print(Correct_answers)
                with open('sample_input/responses.csv','r') as file:
                    reader=csv.reader(file)
                    for row in reader:
                        if counter==0:
                            Header=row
                            counter+=1
                            continue
                        result_marksheet_data[row[6]]={}
                        Right_answer=0
                        Wrong_answer=0
                        Not_attempted=0
                        marks_scored=0
                        response=[]
                        for i in range(7,len(row)):
                            Student_answer=row[i]
                            response.append([Student_answer,Correct_answers[i-7]])
                            if len(Student_answer.strip())==0:
                                Not_attempted+=1
                            elif Student_answer==Correct_answers[i-7]:
                                marks_scored+=positive_marking
                                Right_answer+=1
                            else:
                                marks_scored+=negative_marking
                                Wrong_answer+=1
                        result_marksheet_data[row[6]]["info"]=[row[0],row[1],str(Right_answer*positive_marking)+"/"+str(Total_Questions*positive_marking),row[3],row[4],row[5],str(marks_scored)+"/"+str(positive_marking*Total_Questions),row[6]]              
                        result_marksheet_data[row[6]]["marked_answers"]=response
                        result_marksheet_data[row[6]]["data"]=[["No.",Right_answer,Wrong_answer,Not_attempted,Total_Questions],["Marking",positive_marking,negative_marking,0,""],["Total",Right_answer*positive_marking,Wrong_answer*negative_marking,"",str(marks_scored)+"/"+str(Total_Questions*positive_marking)]]
                        right_ans_str='['+str(Right_answer)+','+str(Wrong_answer)+','+str(Not_attempted)+']'
                        result_marksheet_data[row[6]]["StatusAns"]=[right_ans_str]

                if not os.path.isdir("./marksheets/"):                               
                    os.mkdir("./marksheets/")
                for key, value in result_marksheet_data.items():
                    from openpyxl import Workbook
                    wb = Workbook()
                    sheet = wb.active
                    sheet.title="quiz"
                    img = openpyxl.drawing.image.Image("IITP LOGO.png")
                    img.anchor = 'A1'
                    img.width = 648
                    img.height = 83
                    sheet.add_image(img)
                    sheet.column_dimensions['A'].width=18
                    sheet.column_dimensions['B'].width=18
                    sheet.column_dimensions['C'].width=18
                    sheet.column_dimensions['D'].width=18
                    sheet.column_dimensions['E'].width=18
                    sheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=5)
                    sheet.cell(row=5,column=1,value='Mark Sheet').font=Font(name='Century',size=18,underline='single',color='000000',bold=True)
                    sheet.cell(row=5,column=1).alignment = Alignment(horizontal='center')
                    sheet.cell(row=6,column=1,value='Name:').font=Font(name='Century',size=12,color='000000')
                    sheet.cell(row=6,column=1).alignment=Alignment(horizontal='right')
                    stud_name=value["info"][3]
                    sheet.cell(row=6,column=2,value=stud_name).font=Font(name='Century',size=12,color='000000',bold=True)
                    sheet.cell(row=6,column=4,value='Exam:').font=Font(name='Century',size=12,color='000000')
                    sheet.cell(row=6,column=4).alignment=Alignment(horizontal='right')
                    sheet.cell(row=6,column=5,value='quiz').font=Font(name='Century',size=12,color='000000',bold=True)
                    sheet.cell(row=7,column=1,value='Roll Number:').font=Font(name='Century',size=12,color='000000')
                    sheet.cell(row=7,column=1).alignment=Alignment(horizontal='right')
                    sheet.cell(row=7,column=2,value=key).font=Font(name='Century',size=12,color='000000',bold=True)

                    header_data = ["","Right","Wrong","Not Attempt","Max"]
                    col_no = 1
                    for i in header_data:
                        sheet.cell(row=9,column=col_no,value=i).font=Font(name='Century',size=12,color='000000',bold=True)
                        sheet.cell(row=9,column=col_no).alignment=Alignment(horizontal='center')
                        sheet.cell(row=9,column=col_no,value=i).border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        col_no+=1

                    data=value["data"]
                    row_no = 10
                    for sub_data in data:
                        col_no=1
                        for x in sub_data:
                            if col_no==1:
                                sheet.cell(row=row_no,column=col_no,value=x).font=Font(name='Century',size=12,color='000000',bold=True)
                                sheet.cell(row=row_no,column=col_no).alignment=Alignment(horizontal='center')
                                sheet.cell(row=row_no,column=col_no,value=x).border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                            elif col_no==2:
                                sheet.cell(row=row_no,column=col_no,value=x).font=Font(name='Century',size=12,color='339933')
                                sheet.cell(row=row_no,column=col_no).alignment=Alignment(horizontal='center')
                                sheet.cell(row=row_no,column=col_no,value=x).border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                            elif col_no==3:
                                sheet.cell(row=row_no,column=col_no,value=x).font=Font(name='Century',size=12,color='FF0000')
                                sheet.cell(row=row_no,column=col_no).alignment=Alignment(horizontal='center')
                                sheet.cell(row=row_no,column=col_no,value=x).border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                            else:
                                sheet.cell(row=row_no,column=col_no,value=x).font=Font(name='Century',size=12,color='000000')
                                sheet.cell(row=row_no,column=col_no).alignment=Alignment(horizontal='center')
                                sheet.cell(row=row_no,column=col_no,value=x).border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                            col_no+=1
                        row_no+=1
                    sheet.cell(row=12,column=5).font=Font(name='Century',size=12,color='0000FF') 
                    for i in [1,2,4,5]:
                        col_no=i
                        if i==1 or i==4:
                            response_header = "Student Ans"
                        else:
                            response_header = "Correct Ans"
                        sheet.cell(row=15,column=col_no,value=response_header).font=Font(name='Century',size=12,color='000000',bold=True)
                        sheet.cell(row=15,column=col_no).alignment=Alignment(horizontal='center')
                        sheet.cell(row=15,column=col_no,value=response_header).border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    row_no=16
                    marked_ans=value["marked_answers"]
                    for i in range(0,len(marked_ans)-3):
                        stud_answer=marked_ans[i][0]
                        corr_answer=marked_ans[i][1]
                        if stud_answer==corr_answer:
                            sheet.cell(row=row_no,column=1,value=stud_answer).font=Font(name='Century',size=12,color='339933')
                            sheet.cell(row=row_no,column=1).alignment=Alignment(horizontal='center')
                            sheet.cell(row=row_no,column=1,value=stud_answer).border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        else:
                            sheet.cell(row=row_no,column=1,value=stud_answer).font=Font(name='Century',size=12,color='FF0000')
                            sheet.cell(row=row_no,column=1).alignment=Alignment(horizontal='center')
                            sheet.cell(row=row_no,column=1,value=stud_answer).border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        
                        sheet.cell(row=row_no,column=2,value=stud_answer).font=Font(name='Century',size=12,color='0000FF')
                        sheet.cell(row=row_no,column=2).alignment=Alignment(horizontal='center')
                        sheet.cell(row=row_no,column=2,value=stud_answer).border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        row_no+=1
                    row_no=16
                    for i in range(len(marked_ans)-3,len(marked_ans)):
                        stud_answer=marked_ans[i][0]
                        corr_answer=marked_ans[i][1]
                        if stud_answer==corr_answer:
                            sheet.cell(row=row_no,column=4,value=stud_answer).font=Font(name='Century',size=12,color='339933')
                            sheet.cell(row=row_no,column=4).alignment=Alignment(horizontal='center')
                            sheet.cell(row=row_no,column=4,value=stud_answer).border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        else:
                            sheet.cell(row=row_no,column=4,value=stud_answer).font=Font(name='Century',size=12,color='FF0000')
                            sheet.cell(row=row_no,column=4).alignment=Alignment(horizontal='center')
                            sheet.cell(row=row_no,column=4,value=stud_answer).border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        
                        sheet.cell(row=row_no,column=5,value=stud_answer).font=Font(name='Century',size=12,color='0000FF')
                        sheet.cell(row=row_no,column=5).alignment=Alignment(horizontal='center')
                        sheet.cell(row=row_no,column=5,value=stud_answer).border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        row_no+=1
                    wb.save("./marksheets/"+key+".xlsx")

            absent_roll_nos = {}
            def blank_marksheet():
                with open('sample_input/master_roll.csv','r') as file:
                    counter=0
                    header=[]
                    reader = csv.reader(file)
                    for row in reader:
                        if counter==0:
                            header=row
                            counter+=1
                            continue
                        with open('sample_input/responses.csv','r') as f:
                            count=0
                            headfile=[]
                            reader=csv.reader(f)
                            flag=0
                            for row_response in reader:
                                if count==0:
                                    headfile=row_response
                                    count+=1
                                    continue
                                if row_response[6]==row[0]:
                                    flag=1
                                    break
                            if flag==0:
                                absent_roll_nos[row[0]]=[row[1]]

                from openpyxl import Workbook
                wb=Workbook()
                sheet=wb.active
                sheet.title="quiz"
                for key,value in absent_roll_nos.items():
                    img = openpyxl.drawing.image.Image("IITP LOGO.png")
                    img.anchor = 'A1'
                    img.width = 648
                    img.height = 83
                    sheet.add_image(img)
                    sheet.column_dimensions['A'].width=18
                    sheet.column_dimensions['B'].width=18
                    sheet.column_dimensions['C'].width=18
                    sheet.column_dimensions['D'].width=18
                    sheet.column_dimensions['E'].width=18
                    sheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=5)
                    sheet.cell(row=5,column=1,value='Mark Sheet').font=Font(name='Century',size=18,underline='single',color='000000',bold=True)
                    sheet.cell(row=5,column=1).alignment = Alignment(horizontal='center')
                    sheet.cell(row=6,column=1,value='Name:').font=Font(name='Century',size=12,color='000000')
                    sheet.cell(row=6,column=1).alignment=Alignment(horizontal='right')
                    stud_name=value[0]
                    sheet.cell(row=6,column=2,value=stud_name).font=Font(name='Century',size=12,color='000000',bold=True)
                    sheet.cell(row=6,column=4,value='Exam:').font=Font(name='Century',size=12,color='000000')
                    sheet.cell(row=6,column=4).alignment=Alignment(horizontal='right')
                    sheet.cell(row=6,column=5,value='quiz').font=Font(name='Century',size=12,color='000000',bold=True)
                    sheet.cell(row=7,column=1,value='Roll Number:').font=Font(name='Century',size=12,color='000000')
                    sheet.cell(row=7,column=1).alignment=Alignment(horizontal='right')
                    sheet.cell(row=7,column=2,value=key).font=Font(name='Century',size=12,color='000000',bold=True)

                    sheet.merge_cells(start_row=9, start_column=2, end_row=11, end_column=4)
                    sheet.cell(row=9,column=2,value='ABSENT').font=Font(name='Century',size=20,color='000000',bold=True)
                    sheet.cell(row=9,column=2).alignment = Alignment(horizontal='center')
                    wb.save("./marksheets/"+key+".xlsx")

            def concise_marksheet():
                from openpyxl import Workbook
                wb=Workbook()
                sheet=wb.active
                sheet.title="concise_marksheet"
                Headfile=['Timestamp','Email address','Google_Score','Name','IITP webmail','Phone (10 digit only)',
                        'Score_After_Negative','Roll Number','Unnamed: 7','Unnamed: 8','Unnamed: 9','Unnamed: 10',
                        'Unnamed: 11','Unnamed: 12','Unnamed: 13','Unnamed: 14','Unnamed: 15','Unnamed: 16','Unnamed: 17',
                        'Unnamed: 18','Unnamed: 19','Unnamed: 20','Unnamed: 21','Unnamed: 22','Unnamed: 23','Unnamed: 24',
                        'Unnamed: 25','Unnamed: 26','Unnamed: 27','Unnamed: 28','Unnamed: 29','Unnamed: 30','Unnamed: 31',
                        'Unnamed: 32','Unnamed: 33','Unnamed: 34','statusAns']
                col_no1=1
                for i in Headfile:
                    sheet.cell(row=1,column=col_no1).value=i
                    col_no1+=1
                row_no1=2
                for key,value in result_marksheet_data.items():
                    info = value['info']
                    response = value['marked_answers']
                    status_ans=value['StatusAns']
                    col_no1=1
                    for x in info:
                        sheet.cell(row=row_no1,column=col_no1).value=x
                        col_no1+=1
                    for y in response:
                        sheet.cell(row=row_no1,column=col_no1).value=y[0]
                        col_no1+=1
                    for z in status_ans:
                        sheet.cell(row=row_no1,column=col_no1).value=z
                    row_no1+=1
                for key1,value1 in absent_roll_nos.items():
                    sheet.cell(row=row_no1,column=1).value="ABSENT"
                    sheet.cell(row=row_no1,column=8).value=key1
                    sheet.cell(row=row_no1,column=4).value=value1[0]
                    row_no1+=1
                wb.save("./marksheets/concise_marksheet.xlsx")
            
            individual_roll_result()
            blank_marksheet()
            concise_marksheet()
            success_concise()
        if st.button("Send Email"):
            with open('sample_input/responses.csv','r') as file:
                reader=csv.reader(file)
                counter=0
                header=[]
                for row in reader:
                    if counter==0:
                        header=row
                        counter+=1
                        continue
                    msg=EmailMessage()
                    msg['subject']='Marksheet'
                    msg['From']='Aditya Raj'
                    msg['To']=row[1],row[4]
                    with open('marksheets/'+row[6]+".xlsx",'rb') as f:
                        file_data=f.read()
                        file_name=f.name
                        msg.add_attachment(file_data,maintype="application",subtype="xlsx",filename=file_name)
                    with smtplib.SMTP_SSL('smtp.gmail.com',465) as server:
                        server.login("rishavraj36901@gmail.com","Adit3670")
                        server.send_message(msg)
                    # break                        #please uncomment break so that only 1 email is sent through the above given sender's email and no spamming is there
            send_email()
    else:
        st.subheader("About")
        st.write("Done by Aditya Raj(1901MM05) & Subhadeep Mandal(1901MM33)")
if __name__ == '__main__':
    main()