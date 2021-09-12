import os             
import csv


def output_by_subject():                               #Function which gives the rollno, register_sem, subno and sub_type of all students in different .xlsx files according to the subno
    
    if not os.path.isdir("./output_by_subject/"):      #Checks whether a directory exists or not
        os.mkdir("./output_by_subject/")               #If directory doesn't exist then creates one

    to_write = {}                                      #Dictionary to store required data
    header = []                                        #List which will store the 1st row of input file
    with open('regtable_old.csv', 'r')as file:         #open the file in read mode
        reader = csv.reader(file)                      #reading through file using csv library
        counter = 0
        for row in reader:                             #traversing through the file
            if counter == 0:
                header = row
                counter = counter+1
                continue
            if row[3] not in to_write:                 #checks for subno in to_write dictionary
                to_write[row[3]] = [[header[0], header[1], header[3], header[8]]]    #If subno not present so header list is added to dictionary
            to_write[row[3]].append([row[0], row[1], row[3], row[8]])                #If subno already present so rest data of subno is added to dictionary

    for key, value in to_write.items():                #traversing through data in to_write dictionary
        from openpyxl import Workbook                  #using openpyxl library
        wb = Workbook() 
        sheet = wb.active
        for row_num in range(1, len(value)):               
            row = value[row_num-1]
            for col_num in range(1, 5):
                sheet.cell(row=(row_num),               #Adding data from to_write dictionary to excel file 
                           column=col_num).value = row[col_num-1]
        wb.save("./output_by_subject/"+key+".xlsx")     #Creating .xlsx file


def output_individual_roll():               #Function which gives the rollno, register_sem, subno and sub_type of all students in different .xlsx files according to the rollno
    
    if not os.path.isdir("./output_individual_roll/"):    #Checks whether a directory exists or not
        os.mkdir("./output_individual_roll/")             #If directory doesn't exist then creates one

    to_write = {}                                         #Dictionary to store required data
    header = []                                           #List which will store the 1st row of input file
    with open('regtable_old.csv', 'r')as file:            #open the file in read mode
        reader = csv.reader(file)                         #reading through file using csv library
        counter = 0
        for row in reader:                                #traversing through the file
            if counter == 0:
                header = row
                counter = counter+1
                continue
            if row[0] not in to_write:                    #checks for rollno in to_write dictionary
                to_write[row[0]] = [[header[0], header[1], header[3], header[8]]]        #If rollno not present so header list is added to dictionary    
            to_write[row[0]].append([row[0], row[1], row[3], row[8]])                    #If rollno already present so rest data of subno is added to dictionary

    for key, value in to_write.items():                   #traversing through data in to_write dictionary
        from openpyxl import Workbook                     #using openpyxl library
        wb = Workbook()
        sheet = wb.active
        for row_num in range(1, len(value)):
            row = value[row_num-1]
            for col_num in range(1, 5):
                sheet.cell(row=(row_num),                  #Adding data from to_write dictionary to excel file 
                           column=col_num).value = row[col_num-1]
        wb.save("./output_individual_roll/"+key+".xlsx")   #Creating .xlsx file


output_by_subject()
output_individual_roll()
